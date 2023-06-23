import re
import time
from openpyxl import Workbook, load_workbook
from Conectores_BD import *
from Atributos import Atributo
import array as array
from Brevo import *

class Contacto:
    def __init__(self, nombre=None, email=None, fecha_nacimiento=None, direccion=None, sexo=None, id_contacto=None, atributos:dict={}):
        self.id_contacto = id_contacto
        self.nombre = nombre
        self.email = email
        self.fecha_nacimiento = fecha_nacimiento
        self.direccion = direccion
        self.sexo = sexo
        self.oculto = 1
        self.atributos = atributos   
    def __str__(self):
        return "ID: {}, Nombre: {}, Email: {}, Fecha de nacimiento: {}, Direccion: {}, Sexo: {}, atributos: {}".format(self.id_contacto, self.nombre, self.email, self.fecha_nacimiento, self.direccion, self.sexo, self.atributos)
    def validar_nombre(nombre: str) -> bool:
        if nombre.isalpha():
            return True
        else:
            return False
    def validar_email(email: str) -> bool:
        if re.match(r"[^@]+@[^@]+\.[^@]+", email):  
            return True
        else:  
            return False 
    def validar_fecha_nacimiento(fecha_nacimiento: str) -> bool:
        try:
            fecha_valida = time.strptime(fecha_nacimiento, "%Y/%m/%d")
            return True
        except:
            return False
    def validar_direccion(direccion: str) -> bool:
        direccion_lista = direccion.split(" ")
        if len(direccion_lista) == 2:
            if direccion_lista[0].isalpha() and direccion_lista[1].isnumeric():
                return True
            else:
                return False
    def validar_sexo(sexo: int) -> bool:
        if sexo in [0,1]:
            return True
        else:
            return False
    validar_dicc = {"nombre": validar_nombre,
                    "email": validar_email,
                    "fecha_nacimiento": validar_fecha_nacimiento,
                    "direccion": validar_direccion,
                    "sexo": validar_sexo}
    def validar(self, key: str, data: str) -> bool:
        return self.validar_dicc[key](data)

class Contacto_Controller():
    def agregar_contacto(newContact: Contacto):
        # Metodo para agregar un contacto con id unico

        brevo = Brevo_contactos()
        id_contacto = brevo.post_contacto(newContact)
        newContact.id = id_contacto.id
        Contacto_Model.Post_contacto(newContact)    
    def obtener_contactos_params(habilitado, id_contacto=None, nombre=None, email=None, fecha_nacimiento=None, direccion=None, sexo=None, edad_desde=None, edad_hasta=None, atributo=list):
        # Metodo para imprimir los contactos no ocultos con parametro

        datos = Contacto_Model.listar_contactos_buscador(id_contacto, nombre, email, fecha_nacimiento, direccion, sexo, edad_desde, edad_hasta, atributo, habilitado)
        resultados = []
        for dato in datos:
            contacto_devuelto = Contacto(id_contacto=dato[0], nombre=dato[1], fecha_nacimiento=dato[2], email=dato[3], direccion=dato[4], sexo=dato[5])
            resultados.append(contacto_devuelto)
        return resultados
    def obtener_contactos(habilitado,id_contacto=None, nombre=None, email=None, fecha_nacimiento=None, direccion=None, sexo=None):
        # Metodo para imprimir los contactos con parametro

        datos = Contacto_Model.listar_contactos(id_contacto, nombre, email, fecha_nacimiento, direccion, sexo, habilitado)
        resultados = []
        for dato in datos:
            contacto_devuelto = Contacto(id_contacto=dato[0], nombre=dato[1], fecha_nacimiento=dato[2], email=dato[3], direccion=dato[4], sexo=dato[5])
            resultados.append(contacto_devuelto)
        return resultados
    def ocultar_contacto(email):
        # Metodo para ocultar un contacto
        Contacto_Model.ocultar_contacto(Contacto(email=email))
    def agregar_atributos_contacto(contacto):
        #Metodo para vincular atributos a un contacto

        ids_atributos = contacto.atributos.keys()
        for atributo in ids_atributos:
            Contacto_Model.Post_atributos_contacto(atributo,contacto.id_contacto)
    def eliminar_atributos_contacto(contacto):
        #Metodo para desvincular atributos a un contacto

        ids_atributos = contacto.atributos.keys()
        for atributo in ids_atributos:
            Contacto_Model.Delete_atributos_contacto(atributo,contacto.id_contacto)
    def obtener_contactos_atributos(habilitado):
        #Metodo para obtener contactos con sus atributos

        datos = Contacto_Model.Listar_contactos_atributos()

        resultados = {}

        for dato in datos:
            contacto_devuelto = Contacto(id_contacto=dato[0], nombre=dato[1],email=dato[2], atributos={})
            atributo_devuelto = Atributo(ID_atributo=dato[4], AtributoDescripcion=dato[3])

            if dato[4] != None:
                contacto_devuelto.atributos[atributo_devuelto.ID_atributo] = atributo_devuelto.AtributoDescripcion

            if contacto_devuelto.id_contacto not in resultados:
                resultados[contacto_devuelto.id_contacto] = contacto_devuelto
            else:
                resultados[contacto_devuelto.id_contacto].atributos[atributo_devuelto.ID_atributo] = atributo_devuelto.AtributoDescripcion
                
        return resultados


    ###  Metodos para importacion masiva de contactos
    def importacionMasivaPlantilla(pathToSave):
        # Crear la plantilla de importacion
        #Crear un archivo de Excel con valores en la primera fila
        wb = Workbook()
        ws = wb.active
        ws.append(["Ingrese los datos de los contactos que quiere agregar. Por favor, respete el formato de la plantilla (no elimine esta fila). Limite: 100 contactos por importacion"])
        ws.append(["Nombre", "Email", "Fecha de nacimiento", "Direccion", "Sexo (M-F)"])
        
        wb.save(pathToSave)
    def importacionMasivaProcesamiento(path):
        #Procesar la plantilla de importacion

        #leer las filas de un archivo de excel
        wb = load_workbook(path)
        ws = wb.active
        filasErrores = []
        contadorProcesadas = 0
        x = 3 #Valor de la primera fila con datos
        for row in ws.iter_rows(min_row=x, max_row=103, values_only=True):
            #('torito', 'aviscardi@cotymania.com', datetime.datetime(2001, 4, 11, 0, 0), 'del facon 1025', 'M')

            if row[1] != None:
                sexo = row[4].replace('M','0').replace('F','1').replace('m','0').replace('f','1')
                try:
                    NewContact = Contacto(email=row[1], nombre=row[0],direccion=row[3],sexo=sexo,fecha_nacimiento=row[2])
                    Contacto_Controller.agregar_contacto(NewContact)
                    contadorProcesadas += 1
                except:
                    filasErrores.append(x)
            x += 1
        return filasErrores, contadorProcesadas

class Contacto_Model:
    def Post_contacto(contacto):
        # Metodo para agregar un contacto
        #Conector
        MySql = Conectores_BD.conector_mysql()
        
        #Creo el cursor
        cnxn = MySql.raw_connection()
            
        #Qwery
        qwery = "INSERT INTO CONTACTOS (ID_CONTACTO, NOMBRE_CONTACTO, FECHA_NACIMIENTO, EMAIL, DIRECCION, SEXO) VALUES ('{}','{}','{}','{}','{}','{}')".format(contacto.id,contacto.nombre, contacto.fecha_nacimiento, contacto.email, contacto.direccion, contacto.sexo)

        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()            
    def listar_contactos_buscador(id_contacto, nombre, email, fecha_nacimiento, direccion, sexo, edad_desde, edad_hasta, atributo, habilitado):
        # Metodo para buscar contactos. Es especializado para la creacion de ListaContactos
        #Conector
        MySql = Conectores_BD.conector_mysql()
                
        #Qwery
        qwery = """ SELECT 
                        CONTACTOS.ID_CONTACTO AS ID_CONTACTO,
                        NOMBRE_CONTACTO,
                        FECHA_NACIMIENTO,
                        EMAIL,
                        DIRECCION,
                        SEXO
                    FROM CONTACTOS 
                    WHERE CONTACTO_HABILITACION = {} """.format(habilitado)
                
        
        if id_contacto != None:
            optionalwhere = "AND CONTACTOS.ID_CONTACTO = {} ".format(id_contacto)
            qwery = qwery + optionalwhere
                
        if nombre != None:
            optionalwhere = "AND NOMBRE_CONTACTO = '{}' ".format(nombre)
            qwery = qwery + optionalwhere
        
        if fecha_nacimiento != None:
                optionalwhere = "AND FECHA_NACIMIENTO = '{}' ".format(fecha_nacimiento)
                qwery = qwery + optionalwhere
        
        if email != None:
                optionalwhere = "AND EMAIL = '{}' ".format(email)
                qwery = qwery + optionalwhere
        
        if direccion != None:
                optionalwhere = "AND DIRECCION = '{}' ".format(direccion)
                qwery = qwery + optionalwhere

        if sexo != None:
            sexodigito = 0
            if sexo == "Masculino":
                sexodigito = 0
                optionalwhere = "AND SEXO = {} ".format(sexodigito)
                qwery = qwery + optionalwhere
            elif sexo == "Femenino":
                sexodigito = 1
                optionalwhere = "AND SEXO = {} ".format(sexodigito)
                qwery = qwery + optionalwhere
        
        if edad_desde != None and edad_hasta != None:
                optionalwhere = " AND (SELECT TIMESTAMPDIFF(YEAR, FECHA_NACIMIENTO, CURDATE())) BETWEEN {} AND {} ".format(edad_desde, edad_hasta)
                qwery = qwery + optionalwhere
            
        if atributo != []:
            if len(atributo) == 1:
                optionalwhere = "AND ID_CONTACTO IN (SELECT ID_CONTACTO FROM ATRIBUTOS_CONTACTOS WHERE ID_ATRIBUTO = {})".format(atributo[0])
                qwery = qwery + optionalwhere
            elif len(atributo) > 1:
                optionalwhere = "AND ID_CONTACTO IN (SELECT ID_CONTACTO FROM ATRIBUTOS_CONTACTOS WHERE ID_ATRIBUTO IN {})".format(tuple(atributo))
                qwery = qwery + optionalwhere
                    
        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()
            result = conn.execute(text(qwery))
        return result
    def listar_contactos(id_contacto, nombre, email, fecha_nacimiento, direccion, sexo, habilitado):
        # Metodo para buscar contactos en general, permite varios parametros de busqueda
        #Conector
        MySql = Conectores_BD.conector_mysql()
                
        #Qwery
        qwery = """ SELECT 
                        ID_CONTACTO,
                        NOMBRE_CONTACTO,    
                        FECHA_NACIMIENTO,
                        EMAIL,
                        DIRECCION,
                        SEXO
                    FROM CONTACTOS 
                    WHERE CONTACTO_HABILITACION = {} """.format(habilitado)
                
        if id_contacto != None:
            optionalwhere = "AND ID_CONTACTO = '{}' ".format(id_contacto)
            qwery = qwery + optionalwhere
                
        if nombre != None:
            optionalwhere = "AND NOMBRE_CONTACTO = '{}'".format(nombre)
            qwery = qwery + optionalwhere
        
        if fecha_nacimiento != None:
                optionalwhere = "AND FECHA_NACIMIENTO = '{}'".format(fecha_nacimiento)
                qwery = qwery + optionalwhere
        
        if email != None:
                optionalwhere = "AND EMAIL = '{}'".format(email)
                qwery = qwery + optionalwhere
        
        if direccion != None:
                optionalwhere = "DIRECCION = '{}'".format(direccion)
                qwery = qwery + optionalwhere

        if sexo != None:
                optionalwhere = "AND SEXO = '{}'".format(sexo)
                qwery = qwery + optionalwhere           
        
        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()
            result = conn.execute(text(qwery))
        return result   
    def ocultar_contacto(contacto):
        # Metodo para "Eliminar" un contacto
        #Conector
        MySql = Conectores_BD.conector_mysql()
        
        #Qwery
        qwery = "UPDATE CONTACTOS SET CONTACTO_HABILITACION = 0 WHERE EMAIL = '{}'".format(contacto.email)
        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()          
    def Post_atributos_contacto(id_atributo,id_contacto):
        # Metodo para agregar un atributo a un contacto
        #Conector
        MySql = Conectores_BD.conector_mysql()
        #Qwery
        qwery = "INSERT INTO ATRIBUTOS_CONTACTOS (ID_ATRIBUTO, ID_CONTACTO) VALUES ('{}','{}')".format(id_atributo, id_contacto)

        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()
    def Delete_atributos_contacto(id_atributo,id_contacto):
        # Metodo para eliminar un atributo de un contacto
        #Conector
        MySql = Conectores_BD.conector_mysql()
        #Qwery
        qwery = "DELETE FROM ATRIBUTOS_CONTACTOS WHERE ID_ATRIBUTO = '{}' AND ID_CONTACTO = '{}';".format(id_atributo, id_contacto)

        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()
    def Listar_contactos_atributos(habilitado:int = 1):
        # Metodo para listar todos los contactos con sus atributos
        #Conector
        MySql = Conectores_BD.conector_mysql()
                
        #Qwery
        qwery = """SELECT
                        CONTACTOS.ID_CONTACTO,
                        NOMBRE_CONTACTO,    
                        EMAIL,
                        ATRIBUTO_DESCRIPCION,
                        ATRIBUTOS.ID_ATRIBUTO
                    FROM CONTACTOS
                        LEFT JOIN ATRIBUTOS_CONTACTOS ON
                            CONTACTOS.ID_CONTACTO = ATRIBUTOS_CONTACTOS.ID_CONTACTO
                        LEFT JOIN ATRIBUTOS ON
                            ATRIBUTOS.ID_ATRIBUTO = ATRIBUTOS_CONTACTOS.ID_ATRIBUTO
                    WHERE CONTACTO_HABILITACION = {}
                    ORDER BY ID_CONTACTO;""".format(habilitado)
                        
        #Ejecuto el comando y guardo cambios
        with MySql.connect() as conn:
            conn.execute(text(qwery))
            conn.commit()
            result = conn.execute(text(qwery))
        return result   